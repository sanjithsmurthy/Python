from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
lang=["kannada","telugu","tamil"]
capital=["bengaluru","hyderabad","chennai"]
state=["karnataka","telangana","tamilnadu"]
code=["KA","TS","TN"]
c=input("enter code to find language and capital")
for i in range(1,4):
    sheet.cell(i,1).value=state[i-1]
    sheet.cell(i,2). value = lang[i-1]
    sheet.cell(i,3).value = capital[i-1]
    sheet.cell(i,4). value = code[i-1]
    if c==sheet.cell(i,4).value:
        print(sheet.cell(i,2).value,"and",sheet.cell(i,3).value)
